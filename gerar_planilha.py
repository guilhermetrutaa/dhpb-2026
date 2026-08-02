import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Cenários Diários
data_cenarios = [
    ["Cenário", "Descrição", "Leituras Estimadas", "Limite Free Tier", "Status", "Custo Excedente (US$)", "Custo Excedente (R$)"],
    ["Dia Normal", "Acessos casuais, montagem de equipes, ~500 usuários", "15.000", "50.000", "Coberto (100% Free)", "$0.00", "R$ 0,00"],
    ["Dia de Prova (Média)", "Fases 1 e 2, ~4000 usuários ativos", "136.000", "50.000", "Excede (Requer Blaze)", "$0.03", "R$ 0,17"],
    ["Pico Final", "Fase Final, ~8000 usuários ativos simultâneos", "272.000", "50.000", "Excede (Requer Blaze)", "$0.08", "R$ 0,45"],
    ["Fechamento de Fase", "Acesso Admin, aprovações, painéis (Otimizado)", "5.000", "50.000", "Coberto (100% Free)", "$0.00", "R$ 0,00"]
]

# Projeções Mensais/Anuais (Estimativas)
data_projecoes = [
    ["Período", "Meses", "Descrição", "Custo Estimado (R$)"],
    ["Fase de Cadastro", "Agosto a Setembro", "Acessos normais, dias tranquilos dentro do Spark", "R$ 0,00"],
    ["Mês de Competição", "Outubro", "2 dias de prova média + acessos normais", "R$ 0,34"],
    ["Mês do Pico Final", "Novembro/Dez", "1 dia de pico final + fechamentos e correções", "R$ 0,45"],
    ["Hospedagem Front-end", "Anual", "Hospedagem na Vercel (Gratuita para projetos hobby/edu)", "R$ 0,00"],
    ["Domínio", "Anual", "Domínio institucional (.edu.br ou gratuito)", "R$ 0,00"],
    ["CUSTO ANUAL TOTAL", "12 meses", "Toda a infraestrutura do evento (Banco e Hospedagem)", "R$ 0,79"]
]

wb = Workbook()
ws1 = wb.active
ws1.title = "Cenários Diários"

for row in data_cenarios:
    ws1.append(row)

ws2 = wb.create_sheet(title="Projeção Mensal e Anual")
for row in data_projecoes:
    ws2.append(row)

# Estilização
def style_ws(ws):
    header_fill = PatternFill(start_color="82181A", end_color="82181A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row > 1 and cell.column >= 3:
                cell.alignment = Alignment(horizontal="center")
            
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[col_letter].width = adjusted_width

style_ws(ws1)
style_ws(ws2)

# Destacar a linha de Custo Anual Total
for cell in ws2[ws2.max_row]:
    cell.font = Font(bold=True, color="82181A")
    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

wb.save("DHPB-2026-Relatorio-Custos.xlsx")
print("Planilha gerada com sucesso!")
